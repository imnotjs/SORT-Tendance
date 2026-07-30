"""
SORT-tendance :: XLSX Report Generator
======================================
Generates per-slot, per-date attendance overview XLSX files.

This is the LECTURER-FACING counterpart to pdf_generator.py. It produces
a clean two-sheet workbook (Summary + Detailed Attendance) with NO
stranger information -- strangers are an internal-security concern and
should not appear on the report the lecturer receives.

Output filename pattern (matches the PDF):
    Overview_Class_{SanitizedSubject}_{YYYY-MM-DD}.xlsx
e.g.
    Overview_Class_WebDesign_2026-07-20.xlsx

Workbook structure:
  Sheet 1 -- "Summary"
      Title at B2, metadata block (Subject, Date, Time Slot, Slot ID,
      Total Expected, Attended, Not Attended, Attendance Rate). Uses
      live COUNTIF formulas on the Attendance sheet so the summary
      stays correct if the lecturer edits a status cell.

  Sheet 2 -- "Attendance"
      Title at B2, column headers at row 4 (Student ID, Student Name,
      Status, First Seen, Last Seen), one row per expected student.
      Color-coded rows: green fill for ATTENDED, red fill for
      NOT_ATTENDED. Frozen panes (C5) keep headers + Student ID column
      visible while scrolling. Status column uses data validation
      (dropdown: ATTENDED / NOT ATTENDED) so the lecturer can manually
      override a status if needed.

The generator reads from:
  - storage/attend_csv/attendance_{date}_{slot_id}.csv
(via the same attendance_rows list that pdf_generator.py consumes --
the caller is responsible for loading CSV data via AttendanceEngine.)

Writes to:
  - storage/attendance_reports/Overview_Class_{Subject}_{Date}.xlsx

Author: SORT-tendance Engineering
"""

from __future__ import annotations

import datetime as _dt
import logging
import os
from typing import Any, Dict, List, Optional

logger = logging.getLogger("sortendance.xlsx_generator")

# openpyxl imports -- kept inside try/except so the module imports
# cleanly even on systems without openpyxl (the dashboard will show
# a "XLSX generation unavailable" message instead of crashing).
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    _OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    _OPENPYXL_AVAILABLE = False
    Workbook = None  # type: ignore
    Font = None  # type: ignore
    PatternFill = None  # type: ignore
    Alignment = None  # type: ignore
    Border = Side = None  # type: ignore
    get_column_letter = None  # type: ignore
    DataValidation = None  # type: ignore


# ---------------------------------------------------------------------------
# Design tokens (mirror pdf_generator.py's palette for visual consistency).
# These match the xlsx skill's "professional" palette but are inlined
# here so this module has zero external dependencies on skill
# infrastructure -- it must deploy cleanly to the Windows machine.
# ---------------------------------------------------------------------------
_COLOR_PRIMARY        = "1F3A5F"   # deep blue (header bg, title)
_COLOR_PRIMARY_LIGHT   = "D6E4F0"  # light blue (metadata key cells)
_COLOR_ATTENDED_BG     = "D4EDDA"  # soft green
_COLOR_ATTENDED_FG     = "155724"  # dark green text
_COLOR_NOT_ATTENDED_BG = "F8D7DA"  # soft red
_COLOR_NOT_ATTENDED_FG = "721C24"  # dark red text
_COLOR_NEUTRAL_TEXT    = "37352F"  # body text
_COLOR_NEUTRAL_BORDER  = "CCCCCC"  # cell borders
_COLOR_SUMMARY_BG      = "F1F3F5"  # metadata block bg
_COLOR_WHITE           = "FFFFFF"

# Font name. Microsoft YaHei is the safe default on Windows; Calibri is
# the Excel default. We try YaHei first for CJK compatibility (in case
# student names contain non-Latin characters), and Excel will fall back
# to Calibri automatically if YaHei is unavailable.
_FONT_NAME = "Microsoft YaHei"


# ---------------------------------------------------------------------------
# Public API.
# ---------------------------------------------------------------------------
def generate_attendance_xlsx(
    subject_name: str,
    slot_id: str,
    date_str: str,
    start_time: str,
    end_time: str,
    attendance_rows: List[Dict[str, Any]],
    output_dir: str,
    *,
    include_strangers: bool = False,   # kept for API symmetry; ALWAYS ignored
    stranger_rows: Optional[List[Dict[str, Any]]] = None,
) -> str:
    """Generate a per-slot, per-date attendance overview XLSX.

    This is the lecturer-facing report. Strangers are NEVER included,
    regardless of the ``include_strangers`` flag or ``stranger_rows``
    argument -- those parameters exist only for API symmetry with
    :func:`pdf_generator.generate_attendance_pdf` so the dashboard
    can call them interchangeably.

    Args:
        subject_name: e.g. "AI" or "Web Design".
        slot_id: e.g. "slot_2".
        date_str: e.g. "2026-07-20".
        start_time: e.g. "10:00".
        end_time: e.g. "12:00".
        attendance_rows: List of dicts (one per expected student) with
            keys: student_id, student_name, status, first_seen_in_slot,
            last_seen_in_slot.
        output_dir: Directory to write the XLSX into (created if missing).
        include_strangers: IGNORED. Always False. Kept for API symmetry.
        stranger_rows: IGNORED. Kept for API symmetry.

    Returns:
        Absolute path to the generated XLSX.
    """
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError(
            "openpyxl is not installed. Install with: pip install openpyxl"
        )

    os.makedirs(output_dir, exist_ok=True)
    sanitized = _sanitize_subject(subject_name)
    filename = f"Overview_Class_{sanitized}_{date_str}.xlsx"
    output_path = os.path.join(output_dir, filename)

    wb = Workbook()
    # Sheet 1: Summary.
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _build_summary_sheet(
        ws_sum,
        subject_name=subject_name,
        slot_id=slot_id,
        date_str=date_str,
        start_time=start_time,
        end_time=end_time,
        attendance_rows=attendance_rows,
    )

    # Sheet 2: Detailed Attendance.
    ws_att = wb.create_sheet("Attendance")
    _build_attendance_sheet(
        ws_att,
        subject_name=subject_name,
        date_str=date_str,
        attendance_rows=attendance_rows,
    )

    # Set workbook metadata.
    wb.properties.creator = "SORT-tendance"
    wb.properties.title = (
        f"Class Attendance Overview - {subject_name} - {date_str}"
    )

    wb.save(output_path)
    total = len(attendance_rows)
    attended = sum(1 for r in attendance_rows if r.get("status") == "ATTENDED")
    logger.info(
        "xlsx_generator: wrote %s | attended=%d/%d | strangers=excluded",
        output_path, attended, total,
    )
    return output_path


def is_available() -> bool:
    """Return True if openpyxl is installed and XLSX generation is possible."""
    return _OPENPYXL_AVAILABLE


# ---------------------------------------------------------------------------
# Sheet builders.
# ---------------------------------------------------------------------------
def _build_summary_sheet(
    ws,
    *,
    subject_name: str,
    slot_id: str,
    date_str: str,
    start_time: str,
    end_time: str,
    attendance_rows: List[Dict[str, Any]],
) -> None:
    """Build the Summary sheet (Sheet 1).

    Layout (Canvas Origin B2 per xlsx skill convention):
        B2      : Title (merged B2:E2)
        B3      : Subtitle (generated-at timestamp)
        B5..E8  : Metadata block (4 rows x 4 cols, key/value pairs)
        B10     : Section heading "Attendance Breakdown"
        B12..E13: Counts block (Total / Attended / Not Attended / Rate)
                  with live COUNTIF formulas referencing the Attendance
                  sheet so manual edits to a status cell stay in sync.
    """
    # --- Column widths ---
    ws.column_dimensions["A"].width = 3   # left margin
    ws.column_dimensions["B"].width = 22  # key column
    ws.column_dimensions["C"].width = 28  # value column
    ws.column_dimensions["D"].width = 22  # key column 2
    ws.column_dimensions["E"].width = 28  # value column 2

    # --- Title at B2 ---
    ws.cell(row=2, column=2, value="SORT-tendance :: Class Attendance Overview")
    ws.merge_cells("B2:E2")
    _style_title(ws.cell(row=2, column=2))

    # --- Subtitle at B3 ---
    gen_at = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.cell(row=3, column=2, value=f"Generated {gen_at}")
    ws.merge_cells("B3:E3")
    _style_caption(ws.cell(row=3, column=2))

    # --- Metadata block (B5..E8) ---
    # We need the count of expected students BEFORE building metadata
    # so we can fill the Total Expected cell immediately.
    total = len(attendance_rows)

    meta_rows = [
        ("Subject:",      subject_name,           "Date:",           date_str),
        ("Time Slot:",    f"{start_time} - {end_time}", "Slot ID:",  slot_id),
        ("Total Expected:", str(total),           "Attended:",      ""),  # formula
        ("Not Attended:", "",                                       "Attendance Rate:", ""),  # formulas
    ]
    for row_idx, row_data in enumerate(meta_rows, start=5):
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in (2, 4):  # key cells
                _style_meta_key(cell)
            else:                   # value cells
                _style_meta_value(cell)

    # --- Live formulas referencing the Attendance sheet ---
    # The Attendance sheet has headers at row 4 and data starting at
    # row 5. The Status column is column D (D5:D{4+N}).
    #
    # Metadata block layout (rows 5..8):
    #   B5: Subject:         C5: <subject>     D5: Date:           E5: <date>
    #   B6: Time Slot:       C6: <slot>        D6: Slot ID:        E6: <slot_id>
    #   B7: Total Expected:  C7: <count int>   D7: Attended:       E7: <COUNTIF>
    #   B8: Not Attended:    C8: <C7-E7>       D8: Attendance Rate:E8: <E7/C7>
    n = total
    last_att_row = 4 + n  # row 4 is header; data occupies rows 5..(4+n)
    status_range = f"Attendance!$D$5:$D${last_att_row}"

    # E7 = Attended = COUNTIF(status_range, "ATTENDED")
    ws.cell(row=7, column=5, value=f'=COUNTIF({status_range},"ATTENDED")')
    _style_meta_value(ws.cell(row=7, column=5))
    # C8 = Not Attended = Total (C7) - Attended (E7)
    ws.cell(row=8, column=3, value=f"=C7-E7")
    _style_meta_value(ws.cell(row=8, column=3))
    # E8 = Attendance Rate = IFERROR(Attended / Total, 0)
    rate_cell = ws.cell(row=8, column=5, value=f"=IFERROR(E7/C7,0)")
    rate_cell.number_format = "0.0%"
    _style_meta_value(rate_cell)

    # --- Section: Attendance Breakdown (B10..) ---
    ws.cell(row=10, column=2, value="Attendance Breakdown")
    ws.merge_cells("B10:E10")
    _style_section_heading(ws.cell(row=10, column=2))

    # Mini breakdown table at B12.
    # References the metadata block (rows 5..8):
    #   C7 = Total Expected, E7 = Attended, C8 = Not Attended, E8 = Rate
    breakdown = [
        ("Total Expected",    "=C7"),
        ("Attended",          "=E7"),
        ("Not Attended",      "=C8"),
        ("Attendance Rate",   "=E8"),
    ]
    # Header row at 12.
    ws.cell(row=12, column=2, value="Metric")
    ws.cell(row=12, column=3, value="Value")
    _style_table_header(ws.cell(row=12, column=2))
    _style_table_header(ws.cell(row=12, column=3))
    for i, (label, formula) in enumerate(breakdown, start=13):
        c_label = ws.cell(row=i, column=2, value=label)
        c_value = ws.cell(row=i, column=3, value=formula)
        _style_meta_key(c_label)
        _style_meta_value(c_value)
        if label == "Attendance Rate":
            c_value.number_format = "0.0%"

    # Freeze the title rows.
    ws.freeze_panes = "A5"

    # Print setup.
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True


def _build_attendance_sheet(
    ws,
    *,
    subject_name: str,
    date_str: str,
    attendance_rows: List[Dict[str, Any]],
) -> None:
    """Build the Attendance sheet (Sheet 2).

    Layout:
        B2      : Title (merged B2:F2)
        B3      : Subtitle (subject + date)
        B4..F4  : Column headers (Student ID, Student Name, Status,
                  First Seen, Last Seen)
        B5..    : Data rows (one per expected student), color-coded by status
    """
    # --- Column widths (xlsx skill COLUMN_WIDTHS conventions) ---
    ws.column_dimensions["A"].width = 3    # left margin
    ws.column_dimensions["B"].width = 14   # Student ID
    ws.column_dimensions["C"].width = 28   # Student Name
    ws.column_dimensions["D"].width = 16   # Status
    ws.column_dimensions["E"].width = 14   # First Seen
    ws.column_dimensions["F"].width = 14   # Last Seen

    # --- Title at B2 ---
    ws.cell(row=2, column=2, value="Detailed Attendance")
    ws.merge_cells("B2:F2")
    _style_title(ws.cell(row=2, column=2))

    # --- Subtitle at B3 ---
    ws.cell(row=3, column=2, value=f"{subject_name} — {date_str}")
    ws.merge_cells("B3:F3")
    _style_caption(ws.cell(row=3, column=2))

    # --- Column headers at row 4 ---
    headers = ["Student ID", "Student Name", "Status", "First Seen", "Last Seen"]
    for col_idx, h in enumerate(headers, start=2):
        cell = ws.cell(row=4, column=col_idx, value=h)
        _style_table_header(cell)

    # --- Data rows starting at row 5 ---
    border_thin = _border_thin()
    fill_attended = PatternFill(
        start_color=_COLOR_ATTENDED_BG,
        end_color=_COLOR_ATTENDED_BG,
        fill_type="solid",
    )
    fill_not_attended = PatternFill(
        start_color=_COLOR_NOT_ATTENDED_BG,
        end_color=_COLOR_NOT_ATTENDED_BG,
        fill_type="solid",
    )
    font_attended = Font(
        name=_FONT_NAME, size=10, color=_COLOR_ATTENDED_FG, bold=True,
    )
    font_not_attended = Font(
        name=_FONT_NAME, size=10, color=_COLOR_NOT_ATTENDED_FG, bold=True,
    )
    font_body = Font(name=_FONT_NAME, size=10, color=_COLOR_NEUTRAL_TEXT)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    align_center = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(attendance_rows, start=5):
        sid = str(r.get("student_id", ""))
        sname = str(r.get("student_name", "") or "(unknown)")
        status = "ATTENDED" if r.get("status") == "ATTENDED" else "NOT ATTENDED"
        first_seen = str(r.get("first_seen_in_slot", "") or "—")
        last_seen = str(r.get("last_seen_in_slot", "") or "—")

        # Student ID
        c = ws.cell(row=i, column=2, value=sid)
        c.font = font_body
        c.alignment = align_center
        c.border = border_thin
        # Student Name
        c = ws.cell(row=i, column=3, value=sname)
        c.font = font_body
        c.alignment = align_left
        c.border = border_thin
        # Status (color-coded)
        c = ws.cell(row=i, column=4, value=status)
        c.alignment = align_center
        c.border = border_thin
        if status == "ATTENDED":
            c.fill = fill_attended
            c.font = font_attended
        else:
            c.fill = fill_not_attended
            c.font = font_not_attended
        # First Seen
        c = ws.cell(row=i, column=5, value=first_seen)
        c.font = font_body
        c.alignment = align_center
        c.border = border_thin
        # Last Seen
        c = ws.cell(row=i, column=6, value=last_seen)
        c.font = font_body
        c.alignment = align_center
        c.border = border_thin

    # --- Data validation: Status dropdown (lets lecturer override) ---
    last_row = 4 + len(attendance_rows)
    if len(attendance_rows) > 0:
        dv = DataValidation(
            type="list",
            formula1='"ATTENDED,NOT ATTENDED"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Invalid status",
            error="Status must be ATTENDED or NOT ATTENDED",
        )
        dv.add(f"D5:D{last_row}")
        ws.add_data_validation(dv)

    # --- Conditional formatting on Status column via fills ---
    # (We already applied per-row fills above, but adding CF rules too
    # so the coloring auto-updates if the lecturer changes a status via
    # the dropdown.)
    from openpyxl.formatting.rule import CellIsRule
    if len(attendance_rows) > 0:
        rng = f"D5:D{last_row}"
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="equal",
                formula=['"ATTENDED"'],
                fill=fill_attended,
                font=font_attended,
            ),
        )
        ws.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="equal",
                formula=['"NOT ATTENDED"'],
                fill=fill_not_attended,
                font=font_not_attended,
            ),
        )

    # Freeze headers (row 4) and Student ID column (col B).
    # C5 = freeze everything above row 5 and left of column C.
    ws.freeze_panes = "C5"

    # Print setup.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "4:4"  # Repeat header on each printed page.


# ---------------------------------------------------------------------------
# Style helpers.
# ---------------------------------------------------------------------------
def _style_title(cell) -> None:
    cell.font = Font(name=_FONT_NAME, size=16, bold=True, color=_COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_caption(cell) -> None:
    cell.font = Font(name=_FONT_NAME, size=9, color="6C757D")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_section_heading(cell) -> None:
    cell.font = Font(name=_FONT_NAME, size=13, bold=True, color=_COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(
        start_color=_COLOR_PRIMARY_LIGHT,
        end_color=_COLOR_PRIMARY_LIGHT,
        fill_type="solid",
    )


def _style_meta_key(cell) -> None:
    cell.font = Font(name=_FONT_NAME, size=10, bold=True, color=_COLOR_PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(
        start_color=_COLOR_SUMMARY_BG,
        end_color=_COLOR_SUMMARY_BG,
        fill_type="solid",
    )
    cell.border = _border_thin()


def _style_meta_value(cell) -> None:
    cell.font = Font(name=_FONT_NAME, size=10, color=_COLOR_NEUTRAL_TEXT)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(
        start_color=_COLOR_WHITE,
        end_color=_COLOR_WHITE,
        fill_type="solid",
    )
    cell.border = _border_thin()


def _style_table_header(cell) -> None:
    cell.font = Font(name=_FONT_NAME, size=10, bold=True, color=_COLOR_WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(
        start_color=_COLOR_PRIMARY,
        end_color=_COLOR_PRIMARY,
        fill_type="solid",
    )
    cell.border = _border_thin()


def _border_thin():
    side = Side(style="thin", color=_COLOR_NEUTRAL_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


# ---------------------------------------------------------------------------
# Helpers.
# ---------------------------------------------------------------------------
def _sanitize_subject(name: str) -> str:
    """Filesystem-safe subject name. 'Web Design' -> 'WebDesign'."""
    out = []
    for ch in str(name):
        if ch.isalnum():
            out.append(ch)
    return "".join(out) or "Unknown"
